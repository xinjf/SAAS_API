from openpyxl import load_workbook

from lib.generate_logs import warning
from utils.ramdom_params import RandomParams
from utils.settings import BASE_PATH, operator_url



class OperateExcel:
    """
    file_name: 读取的excel名称
    sheet_name: 读取的excel的sheet栏[-
    """

    def __init__(self, file_name=None, sheet_name=None):
        self.file_path = BASE_PATH + file_name
        self.sheet_name = sheet_name

    def read_excel_data(self):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2, sheet.max_row + 1):
            sub_data = {
                "case_id": sheet.cell(i, 1).value,
                "detail": sheet.cell(i, 2).value,
                "method": sheet.cell(i, 3).value,
                "header": sheet.cell(i, 4).value,
                "url": sheet.cell(i, 5).value,
                "data": sheet.cell(i, 6).value,
                "sql": sheet.cell(i, 7).value,
                "sql_result": sheet.cell(i, 8).value,
                "check_result": sheet.cell(i, 9).value}
            if sub_data["case_id"] is not None:
                if sub_data["method"]=="sql":
                    test_data.append(sub_data)
                else:
                    try:
                        sub_data["url"] = operator_url + sub_data["url"]
                        sub_data["check_result"] = eval(sub_data["check_result"])
                        sub_data["data"] = RandomParams().build_random_params(sub_data["data"])
                        test_data.append(sub_data)
                    except:
                        warning("接口参数有误，请检查参数格式")
        wb.close()  # 读取文件后需要关闭，否则会报无打开文件权限
        return test_data

    def write_excel_data(self, i, result, Test_result):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        sheet.cell(i, 10).value = result
        sheet.cell(i, 11).value = Test_result
        wb.save(filename=self.file_path)
        wb.close()

    def get_excel_sheet(self):
        wb = load_workbook(self.file_path)
        sheet_list = wb.sheetnames
        return sheet_list


if __name__ == "__main__":
    OperateExcel(r"\test_data\AssetManagement\common\common.xlsx", sheet_name="Store").read_excel_data()
